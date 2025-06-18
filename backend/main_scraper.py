#!/usr/bin/env python3
"""
Replacement main.py for GoogleSheetsProcessor with improved fraction handling
- Puts Title, Description, Price after Mfr Model
- Properly handles fractions in sup/sub tags
- Extracts large product images
"""

import sys
import os
import pandas as pd
import gspread
import re
import math
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton,
    QProgressBar, QScrollArea, QFrame, QMessageBox, QComboBox
)
from PyQt5.QtCore import Qt, QTimer, pyqtSignal, QObject
from PyQt5.QtGui import QFont
from oauth2client.service_account import ServiceAccountCredentials
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from fake_useragent import UserAgent
import threading
import time
import traceback
import openpyxl
from openpyxl.styles import Alignment
import json
import requests
from io import BytesIO
from PIL import Image

# Simple class for better error handling
class AppError(Exception):
    pass

# Define a signal class for thread-safe GUI updates
class WorkerSignals(QObject):
    update_progress = pyqtSignal(int, int)
    update_status = pyqtSignal(str)
    finished = pyqtSignal()
    error = pyqtSignal(str)

class SheetRow(QFrame):
    def __init__(self, index, parent):
        super().__init__(parent)
        self.index = index
        self.parent = parent
        self.running = False
        self.completed = False
        self.output_df = None
        self.output_path = None
        self.selected_file = None
        self.worker_thread = None
        self.signals = WorkerSignals()
        
        # Set up UI
        self.setup_ui()
        
        # Connect signals for thread-safe updates
        self.signals.update_progress.connect(self.on_update_progress)
        self.signals.update_status.connect(self.on_update_status)
        self.signals.finished.connect(self.on_processing_finished)
        self.signals.error.connect(self.on_processing_error)
        
        # Load files in dropdown
        QTimer.singleShot(500, self.load_files)
    
    def setup_ui(self):
        # Basic styling
        self.setFrameShape(QFrame.StyledPanel)
        self.setStyleSheet("""
            QFrame {
                background-color: white;
                border-radius: 2px;
                padding: 8px;
                border: 1px solid #cccccc;
            }
            QProgressBar {
                border: 1px solid #cccccc;
                border-radius: 2px;
                text-align: center;
                height: 14px;
                font-size: 9px;
                margin-top: 0px;
            }
            QProgressBar::chunk {
                background-color: #4285f4;
                border-radius: 1px;
            }
            QComboBox {
                border: 1px solid #aaaaaa;
                border-radius: 2px;
                padding: 3px;
                background-color: white;
                min-height: 24px;
                max-height: 24px;
                font-size: 11px;
            }
            QLineEdit {
                border: 1px solid #aaaaaa;
                border-radius: 2px;
                padding: 3px;
                background-color: white;
                min-height: 24px;
                max-height: 24px;
                font-size: 11px;
            }
            QPushButton {
                border-radius: 2px;
                padding: 3px;
                min-height: 24px;
                max-height: 24px;
                font-weight: bold;
                font-size: 11px;
            }
            QPushButton#startBtn {
                background-color: #4285f4;
                color: white;
                border: none;
            }
            QPushButton#startBtn:hover {
                background-color: #3367d6;
            }
            QPushButton#startBtn:disabled {
                background-color: #a5c2f5;
            }
            QPushButton#stopBtn {
                background-color: #f0f0f0;
                color: #333333;
                border: 1px solid #cccccc;
            }
            QPushButton#stopBtn:hover {
                background-color: #e0e0e0;
            }
            QPushButton#stopBtn:disabled {
                color: #aaaaaa;
                border: 1px solid #e0e0e0;
            }
            QPushButton#refreshBtn {
                background-color: #f0f0f0;
                color: #333333;
                border: 1px solid #aaaaaa;
                font-size: 13px;
                min-width: 24px;
                max-width: 24px;
            }
            QPushButton#refreshBtn:hover {
                background-color: #e0e0e0;
            }
            QLabel {
                color: #333333;
                font-size: 11px;
            }
        """)
        
        # Set object name for debugging
        self.setObjectName(f"SheetRow_{self.index}")
        
        # Main layout
        layout = QVBoxLayout(self)
        layout.setSpacing(6)
        layout.setContentsMargins(10, 8, 10, 8)
        
        # Top row: file selection and buttons
        top_row = QHBoxLayout()
        top_row.setSpacing(4)
        
        # File dropdown
        self.file_dropdown = QComboBox(self)
        self.file_dropdown.setFixedHeight(24)
        self.file_dropdown.addItem("")
        self.file_dropdown.setPlaceholderText("Select File")
        
        # Connect selection change event
        self.file_dropdown.currentIndexChanged.connect(self.file_selected)
        
        # Refresh button
        self.refresh_btn = QPushButton("↻", self)
        self.refresh_btn.setObjectName("refreshBtn")
        self.refresh_btn.setFixedSize(24, 24)
        self.refresh_btn.clicked.connect(self.load_files)
        
        # Prefix input
        self.prefix_input = QLineEdit(self)
        self.prefix_input.setPlaceholderText("Prefix")
        self.prefix_input.setFixedSize(100, 24)
        
        # Start button
        self.start_btn = QPushButton("Start", self)
        self.start_btn.setObjectName("startBtn")
        self.start_btn.setFixedSize(80, 24)
        self.start_btn.clicked.connect(self.start_processing)
        
        # Stop button
        self.stop_btn = QPushButton("Stop", self)
        self.stop_btn.setObjectName("stopBtn")
        self.stop_btn.setFixedSize(80, 24)
        self.stop_btn.clicked.connect(self.stop_processing)
        self.stop_btn.setEnabled(False)
        
        # Add widgets to top row
        top_row.addWidget(self.file_dropdown, 1)
        top_row.addWidget(self.refresh_btn)
        top_row.addSpacing(4)
        top_row.addWidget(self.prefix_input)
        top_row.addSpacing(4)
        top_row.addWidget(self.start_btn)
        top_row.addSpacing(4)
        top_row.addWidget(self.stop_btn)
        
        # Middle row: Status information
        status_row = QHBoxLayout()
        status_row.setContentsMargins(0, 3, 0, 3)
        
        self.status_label_prefix = QLabel("Status:", self)
        self.status_label_prefix.setFont(QFont("Arial", 8, QFont.Bold))
        
        self.status_label = QLabel("Ready", self)
        
        status_row.addWidget(self.status_label_prefix)
        status_row.addWidget(self.status_label)
        status_row.addStretch(1)
        
        # Bottom row: Progress bar
        progress_row = QHBoxLayout()
        
        self.progress_bar = QProgressBar(self)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setFixedHeight(14)
        
        progress_row.addWidget(self.progress_bar)
        
        # Add layouts to main layout
        layout.addLayout(top_row)
        layout.addLayout(status_row)
        layout.addLayout(progress_row)
    
    def on_update_progress(self, current, total):
        percent = 0 if total <= 0 else int((current / total) * 100)
        self.progress_bar.setValue(percent)
        selected_file = self.get_selected_file()
        if selected_file:
            self.parent.update_status(f"Processing: {selected_file['name']} - Row {current} of {total}")
    
    def on_update_status(self, status_text):
        self.status_label.setText(status_text)
    
    def on_processing_finished(self):
        self.running = False
        self.completed = True
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.status_label.setText("Completed")
        selected_file = self.get_selected_file()
        if selected_file:
            self.parent.update_status(f"Completed: {selected_file['name']}")
        self.parent.process_next_row()
    
    def on_processing_error(self, error_message):
        self.running = False
        self.completed = False
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.status_label.setText(f"Error: {error_message[:40]}...")
        QMessageBox.warning(self, "Processing Error", error_message)
        self.parent.process_next_row()  # Continue to next file
    
    def lock_controls(self, locked=True):
        self.file_dropdown.setEnabled(not locked)
        self.prefix_input.setEnabled(not locked)
        self.refresh_btn.setEnabled(not locked)
    
    def load_files(self):
        try:
            self.file_dropdown.currentIndexChanged.disconnect(self.file_selected)
        except:
            pass
        current_selection = self.file_dropdown.currentText()
        self.file_dropdown.clear()
        self.file_dropdown.addItem("")
        self.status_label.setText("Loading files...")
        try:
            all_files = self.parent.get_drive_web_files()
            selected_files = self.parent.get_selected_files()
            available_files = [fn for fn in all_files if fn not in selected_files or fn == current_selection]
            for file_name in available_files:
                self.file_dropdown.addItem(file_name)
            self.status_label.setText(f"Found {len(available_files)} files" if available_files else "No files available")
            if current_selection:
                index = self.file_dropdown.findText(current_selection)
                if index >= 0:
                    self.file_dropdown.setCurrentIndex(index)
        except Exception as e:
            print(f"Error loading files: {e}")
            self.status_label.setText("Error loading files")
        self.file_dropdown.currentIndexChanged.connect(self.file_selected)
    
    def file_selected(self):
        file_name = self.file_dropdown.currentText()
        self.selected_file = file_name if file_name else None
        if file_name:
            self.extract_prefix_from_filename(file_name)
        for i in range(self.parent.scroll_layout.count()):
            item = self.parent.scroll_layout.itemAt(i)
            if item:
                row = item.widget()
                if row and row != self and not row.running:
                    row.load_files()
    
    def extract_prefix_from_filename(self, filename):
        match = re.search(r'[\w]+-(\d+)', filename)
        if match:
            self.prefix_input.setText(match.group(1))
    
    def get_selected_file(self):
        text = self.file_dropdown.currentText()
        if not text:
            return None
        path = os.path.expanduser(f"~/GoogleDriveMount/Web/{text}")
        return {"name": text, "path": path, "type": "local_file"}
    
    def start_processing(self):
        if self.running:
            return
        if not self.file_dropdown.currentText():
            QMessageBox.warning(self, "Error", "Please select a file first")
            return
        prefix = self.prefix_input.text().strip()
        if not prefix:
            QMessageBox.warning(self, "Error", "Please enter a Katom prefix")
            return
        self.running = True
        self.completed = False
        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.progress_bar.setValue(0)
        self.status_label.setText("Starting...")
        self.worker_thread = threading.Thread(target=self.process_file)
        self.worker_thread.daemon = True
        self.worker_thread.start()
    
    def stop_processing(self):
        if not self.running:
            return
        self.running = False
        self.completed = True
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.status_label.setText("Stopped")
    
    def reset_state(self):
        self.running = False
        self.completed = False
        self.progress_bar.setValue(0)
        self.status_label.setText("Ready")
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
    
    def process_weight_value(self, value):
        try:
            number_match = re.search(r'(\d+(\.\d+)?)', str(value))
            if number_match:
                number = float(number_match.group(1))
                rounded = math.ceil(number)
                final = rounded + 5
                units_match = re.search(r'[^\d.]+$', str(value))
                units = units_match.group(0).strip() if units_match else ""
                return f"{final}{' ' + units if units else ''}"
            return value
        except:
            return value

    def check_image_size(self, image_url):
        """Check if an image is larger than 300x300 pixels"""
        try:
            response = requests.get(image_url, timeout=10)
            if response.status_code != 200:
                return False
            
            img = Image.open(BytesIO(response.content))
            width, height = img.size
            
            return width >= 300 and height >= 300
        except Exception as e:
            print(f"Error checking image size for {image_url}: {e}")
            return False
    
    def extract_numeric_price(self, price_text):
        """Extract numeric price value from price text"""
        if not price_text:
            return ""
            
        # Try to find numeric price with decimal point
        price_match = re.search(r'[\d,]+\.\d{2}', price_text)
        if price_match:
            return price_match.group(0).replace(',', '')
            
        # Try to find numeric price without decimal
        price_match = re.search(r'[\d,]+', price_text)
        if price_match:
            return price_match.group(0).replace(',', '')
            
        return ""
    
    def extract_table_data(self, driver):
        specs_dict = {}
        specs_html = ""
        try:
            specs_tables = driver.find_elements(By.CSS_SELECTOR, "table.table.table-condensed.specs-table")
            if not specs_tables:
                specs_tables = driver.find_elements(By.TAG_NAME, "table")
            if specs_tables:
                table = specs_tables[0]
                rows = table.find_elements(By.TAG_NAME, "tr")
                specs_html = '<table class="specs-table" cellspacing="0" cellpadding="4" border="1" style="margin-top:10px;border-collapse:collapse;width:auto;" align="left"><tbody>'
                for row in rows:
                    cells = row.find_elements(By.TAG_NAME, "td")
                    if len(cells) >= 2:
                        key = cells[0].text.strip()
                        # Get the cell HTML to preserve fractions
                        cell_html = cells[1].get_attribute('innerHTML')
                        value_text = cells[1].text.strip()
                        
                        # Remove any images from the cell HTML
                        cell_html = re.sub(r'<img[^>]*>', '', cell_html)
                        
                        # For specs_dict, use the text value but process weight specifically
                        if "weight" in key.lower():
                            value = self.process_weight_value(value_text)
                        else:
                            value = value_text
                            
                        # Store in specs dictionary
                        if key and key.lower() not in specs_dict:
                            specs_dict[key.lower()] = value
                        
                        # Use original HTML in specs table to preserve formatting
                        specs_html += f'<tr><td style="padding:3px 8px;"><b>{key}</b></td><td style="padding:3px 8px;">{cell_html}</td></tr>'
                
                specs_html += "</tbody></table>"
                
            # If no specs table found, try other elements
            if not specs_html:
                other_specs = []
                spec_rows = driver.find_elements(By.CSS_SELECTOR, ".specs-row, [class*='spec']")
                if spec_rows:
                    for row in spec_rows:
                        key_elem = row.find_elements(By.CSS_SELECTOR, ".spec-key, .spec-name, [class*='key'], [class*='name']")
                        val_elem = row.find_elements(By.CSS_SELECTOR, ".spec-value, .spec-val, [class*='value'], [class*='val']")
                        if key_elem and val_elem:
                            key = key_elem[0].text.strip()
                            value_html = val_elem[0].get_attribute('innerHTML')
                            value_text = val_elem[0].text.strip()
                            
                            # Remove any images from the value HTML
                            value_html = re.sub(r'<img[^>]*>', '', value_html)
                            
                            # Process weight values
                            if "weight" in key.lower():
                                value = self.process_weight_value(value_text)
                            else:
                                value = value_text
                                
                            # Store in dictionary
                            if key and key.lower() not in specs_dict:
                                specs_dict[key.lower()] = value
                                other_specs.append((key, value_html))
                                
                # Create HTML table from other specs
                if other_specs:
                    specs_html = '<table class="specs-table" cellspacing="0" cellpadding="4" border="1" style="margin-top:10px;border-collapse:collapse;width:auto;" align="left"><tbody>'
                    for key, value_html in other_specs:
                        specs_html += f'<tr><td style="padding:3px 8px;"><b>{key}</b></td><td style="padding:3px 8px;">{value_html}</td></tr>'
                    specs_html += "</tbody></table>"
                    
        except Exception as e:
            print(f"Error extracting table data: {e}")
            print(traceback.format_exc())
            
        return specs_dict, specs_html
    
    def scrape_katom(self, model_number, prefix, retries=2):
        model_number = ''.join(e for e in model_number if e.isalnum()).upper()
        if model_number.endswith("HC"):
            model_number = model_number[:-2]
        url = f"https://www.katom.com/{prefix}-{model_number}.html"
        
        options = Options()
        options.add_argument('--headless')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument(f'user-agent={UserAgent().random}')
        
        driver = None
        title, description = "Title not found", "Description not found"
        specs_data = {}
        specs_html = ""
        video_links = ""
        numeric_price = ""
        main_image = ""
        additional_images = []
        item_found = False
        
        try:
            driver = webdriver.Chrome(options=options)
            driver.set_page_load_timeout(30)
            driver.get(url)
            
            if "404" in driver.title or "not found" in driver.title.lower():
                return title, description, specs_data, specs_html, video_links, numeric_price, main_image, additional_images
            
            # Extract title
            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "h1.product-name.mb-0, h1"))
                )
                title_element = driver.find_element(By.CSS_SELECTOR, "h1.product-name.mb-0")
                title = title_element.text.strip()
                if title:
                    item_found = True
            except TimeoutException:
                # Try alternate title selectors
                try:
                    for selector in ["h1.product-title", "h1[itemprop='name']", "h1"]:
                        elements = driver.find_elements(By.CSS_SELECTOR, selector)
                        if elements:
                            title = elements[0].text.strip()
                            item_found = True
                            break
                except Exception as e:
                    print(f"Error with alternate title search: {e}")
            except Exception as e:
                print(f"Error getting title: {e}")
            
            if item_found:
                # Extract price - First look specifically for price in <p class="product-price-text m-0">
                try:
                    price_elements = driver.find_elements(By.CSS_SELECTOR, "p.product-price-text.m-0")
                    if price_elements:
                        for element in price_elements:
                            price_text = element.text.strip()
                            if price_text:
                                numeric_price = self.extract_numeric_price(price_text)
                                if numeric_price:
                                    break
                    
                    # If price not found in the specific element, try other selectors
                    if not numeric_price:
                        for selector in [".product-price", ".price", "[class*='price']", ".regular-price", 
                                      ".our-price", ".sale-price", "span[itemprop='price']"]:
                            elements = driver.find_elements(By.CSS_SELECTOR, selector)
                            if elements:
                                for element in elements:
                                    price_text = element.text.strip()
                                    if price_text and ('$' in price_text or re.search(r'\d+\.\d{2}', price_text)):
                                        numeric_price = self.extract_numeric_price(price_text)
                                        if numeric_price:
                                            break
                                if numeric_price:
                                    break
                    
                except Exception as e:
                    print(f"Error extracting price: {e}")
                
                # Extract main image
                try:
                    for selector in [".product-img img", ".main-product-image", "img.main-image", 
                                  "img[itemprop='image']", ".product-image-container img"]:
                        elements = driver.find_elements(By.CSS_SELECTOR, selector)
                        if elements:
                            for element in elements:
                                src = element.get_attribute("src")
                                if src and self.check_image_size(src):
                                    main_image = src
                                    break
                            if main_image:
                                break
                    
                    if not main_image:
                        elements = driver.find_elements(By.TAG_NAME, "img")
                        for element in elements:
                            src = element.get_attribute("src")
                            if not src:
                                continue
                            if (model_number.lower() in src.lower() or "product" in src.lower()) and self.check_image_size(src):
                                main_image = src
                                break
                except Exception as e:
                    print(f"Error extracting main image: {e}")
                
                # Extract additional images
                try:
                    for selector in [".additional-images img", ".product-thumbnails img", ".thumb-image", 
                                  ".product-gallery img", "[class*='thumbnail'] img"]:
                        elements = driver.find_elements(By.CSS_SELECTOR, selector)
                        if elements:
                            for element in elements:
                                src = element.get_attribute("src")
                                # Get the higher resolution version of the image if it's a thumbnail
                                full_size_src = src
                                if src and "thumbnail" in src.lower():
                                    full_size_src = src.replace("thumbnail", "full")
                                
                                if full_size_src and full_size_src != main_image and full_size_src not in additional_images:
                                    if self.check_image_size(full_size_src):
                                        additional_images.append(full_size_src)
                                        if len(additional_images) >= 5:  # Limit to 5 additional images
                                            break
                            if len(additional_images) >= 5:
                                break
                except Exception as e:
                    print(f"Error extracting additional images: {e}")
                
                # Get description
                try:
                    tab_content = driver.find_element(By.CLASS_NAME, "tab-content")
                    paragraphs = tab_content.find_elements(By.TAG_NAME, "p")
                    
                    # The only change from original code is to get innerHTML instead of text
                    # to preserve fraction formatting, but filtering the same way
                    filtered = []
                    for p in paragraphs:
                        p_html = p.get_attribute('innerHTML').strip()
                        p_text = p.text.strip()  # For filtering logic - exactly as original
                        
                        # Filter out paragraphs with conditions from original code
                        if p_text and not p_text.lower().startswith("*free") and "video" not in p_text.lower():
                            # Additional step: remove any img tags from the HTML
                            p_html = re.sub(r'<img[^>]*>', '', p_html)
                            filtered.append(f"<p>{p_html}</p>")
                    
                    description = "".join(filtered) if filtered else "Description not found"
                except NoSuchElementException:
                    try:
                        # Try alternative description selectors - same as original
                        for selector in [".product-description", ".description", 
                                      "[class*='description']", "#product-description", "#description"]:
                            elements = driver.find_elements(By.CSS_SELECTOR, selector)
                            if elements:
                                # Only change from original: use innerHTML instead of text
                                element_html = elements[0].get_attribute('innerHTML')
                                if element_html:
                                    # Remove any img tags
                                    element_html = re.sub(r'<img[^>]*>', '', element_html)
                                    description = f"<p>{element_html}</p>"
                                    break
                    except Exception as e:
                        print(f"Error getting alternate description: {e}")
                except Exception as e:
                    print(f"Error getting description: {e}")
                
                # Extract table data with improved fraction handling
                specs_data, specs_html = self.extract_table_data(driver)
                
                # Extract video links
                try:
                    sources = driver.find_elements(By.CSS_SELECTOR, "source[src*='.mp4'], source[type*='video']")
                    for source in sources:
                        src = source.get_attribute("src")
                        if src and src not in video_links:
                            video_links += f"{src}\n"
                            
                    if not video_links:
                        videos = driver.find_elements(By.TAG_NAME, "video")
                        for video in videos:
                            inner_sources = video.find_elements(By.TAG_NAME, "source")
                            for source in inner_sources:
                                src = source.get_attribute("src")
                                if src and src not in video_links:
                                    video_links += f"{src}\n"
                except Exception as e:
                    print(f"Error extracting video links: {e}")
        except Exception as e:
            print(f"Error in scrape_katom: {e}")
            print(traceback.format_exc())
            if retries > 0:
                if driver:
                    try:
                        driver.quit()
                    except:
                        pass
                time.sleep(2)
                return self.scrape_katom(model_number, prefix, retries - 1)
        finally:
            if driver:
                try:
                    driver.quit()
                except:
                    pass
        
        return title, description, specs_data, specs_html, video_links, numeric_price, main_image, additional_images

    def load_file_data(self, file_info):
        try:
            path = file_info['path']
            
            if path.lower().endswith('.csv'):
                df = pd.read_csv(path)
                print(f"Loaded CSV with {len(df)} rows")
                return df
                
            elif path.lower().endswith(('.xlsx', '.xls')):
                try:
                    df = pd.read_excel(path, engine='openpyxl')
                    print(f"Loaded Excel with {len(df)} rows")
                    return df
                except Exception as e:
                    print(f"Error with openpyxl: {e}")
                    df = pd.read_excel(path)
                    print(f"Loaded Excel with default engine, {len(df)} rows")
                    return df
            else:
                raise AppError(f"Unsupported file type: {path}")
        except Exception as e:
            print(f"Error reading file {path}: {e}")
            raise AppError(f"Failed to read file: {str(e)}")

    def process_file(self):
        """Process the selected file"""
        try:
            file_info = self.get_selected_file()
            if not file_info:
                self.signals.error.emit("No file selected")
                return
            
            prefix = self.prefix_input.text().strip()
            print(f"Starting processing for file: {file_info['path']} with prefix: {prefix}")
            
            # Load data
            try:
                df = self.load_file_data(file_info)
                print(f"Loaded dataframe with columns: {list(df.columns)}")
            except Exception as e:
                self.signals.error.emit(f"Failed to load file: {str(e)}")
                return
            
            # Find model column
            model_col = None
            for col in df.columns:
                col_str = str(col).strip()
                if col_str.lower() == 'mfr model':
                    model_col = col
                    print(f"Found exact model column: {col}")
                    break
            
            # If exact match not found, try partial match
            if not model_col:
                for col in df.columns:
                    col_str = str(col).strip()
                    if 'mfr' in col_str.lower() and 'model' in col_str.lower():
                        model_col = col
                        print(f"Using column '{col}' as model column (partial match)")
                        break
            
            if not model_col:
                # Show all columns to help diagnose issues
                print("Available columns: ", list(df.columns))
                self.signals.error.emit("Missing 'Mfr Model' column in file")
                return
            
            # Define column order with Title, Description, Price right after Mfr Model
            columns = ["Mfr Model", "Title", "Description", "Price"]
            
            # Add image columns
            columns.extend(["Main Image"])
            for i in range(1, 6):  # Up to 5 additional large images
                columns.append(f"Additional Image {i}")
            
            # Add specification columns
            columns.extend(["Manufacturer", "Food Type", "Frypot Style", "Heat", "Hertz", "Nema", 
                           "Number Of Fry Pots", "Oil Capacity/Fryer (Lb)", "Phase", "Product",
                           "Product Type", "Rating", "Special Features", "Type", "Voltage",
                           "Warranty", "Weight", "Dimensions", "Sku", "Shipping Weight"])
            
            # Add video links
            for i in range(1, 6):
                columns.append(f"Video Link {i}")
            
            # Initialize DataFrame with columns
            self.output_df = pd.DataFrame(columns=columns)
            
            # Set up output path - ensure .xlsx extension
            base_name = os.path.splitext(file_info['name'])[0]
            output_name = f"final_{prefix}_{base_name}.xlsx"  # Force .xlsx extension
            
            # Make sure the output directory exists
            output_dir = os.path.expanduser("~/GoogleDriveMount/Web/Completed/Final/")
            os.makedirs(output_dir, exist_ok=True)
            
            self.output_path = os.path.join(output_dir, output_name)
            print(f"Output will be saved to: {self.output_path}")
            
            # Create initial empty file
            try:
                empty_df = pd.DataFrame(columns=columns)
                empty_df.to_excel(self.output_path, index=False)
                print(f"Created initial empty Excel file: {self.output_path}")
            except Exception as e:
                print(f"Warning: Failed to create initial file: {e}")
            
            # Process rows
            total_rows = len(df)
            if total_rows == 0:
                self.signals.error.emit("File contains no data rows")
                return
                
            self.signals.update_progress.emit(0, total_rows)
            print(f"Processing {total_rows} rows")
            
            # We'll collect rows in a list first, then add to DataFrame
            all_rows = []
            processed_count = 0
            
            for i, row_data in df.iterrows():
                if not self.running:
                    break
                    
                current_row = i + 1
                
                # Get model with error handling
                try:
                    model = str(row_data[model_col]).strip()
                    if model.lower() == 'nan' or model.lower() == 'none':
                        model = ""
                except Exception as e:
                    print(f"Error reading model in row {current_row}: {e}")
                    model = ""
                
                if not model:
                    print(f"Skipping row {current_row} - empty model")
                    self.signals.update_progress.emit(current_row, total_rows)
                    continue
                    
                try:
                    self.signals.update_status.emit(f"Processing model: {model}")
                    print(f"Processing row {current_row}: {model}")
                    
                    # Scrape data
                    title, desc, specs_dict, specs_html, video_links, numeric_price, main_image, additional_images = self.scrape_katom(model, prefix)
                    
                    if title != "Title not found" and "not found" not in title.lower():
                        # Use the price as is (just numeric value)
                        price_value = "Call for Price"
                        if numeric_price:
                            price_value = numeric_price
                        
                        combined_description = f'<div style="text-align: justify;">{desc}</div>'
                        if specs_html:
                            combined_description += f'<h3 style="margin-top: 15px;">Specifications</h3>{specs_html}'
                            
                        # Create a dictionary for this row
                        new_row = {}
                        
                        # Initialize all columns to empty string
                        for col in columns:
                            new_row[col] = ""
                            
                        # Set the main values
                        new_row["Mfr Model"] = model
                        new_row["Title"] = title
                        new_row["Description"] = combined_description
                        new_row["Price"] = price_value
                        new_row["Main Image"] = main_image
                        
                        # Add additional images
                        for idx, img_url in enumerate(additional_images[:5], 1):
                            new_row[f"Additional Image {idx}"] = img_url
                                
                        # Add specification data
                        for key, value in specs_dict.items():
                            for field in columns:
                                if key.lower() == field.lower() or key.lower() in field.lower():
                                    new_row[field] = value
                                    break
                                    
                        # Add video links
                        video_list = [link.strip() for link in video_links.strip().split('\n') if link.strip()]
                        for idx, link in enumerate(video_list[:5], 1):
                            new_row[f"Video Link {idx}"] = link
                            
                        # Add to our list
                        all_rows.append(new_row)
                        processed_count += 1
                        
                        # Save periodically
                        if processed_count % 5 == 0 or processed_count == 1:
                            # Create DataFrame from collected rows 
                            self.output_df = pd.DataFrame(all_rows, columns=columns)
                            print(f"Saving after processing {processed_count} rows")
                            self.save_results()
                        
                except Exception as e:
                    print(f"Error processing row {current_row}: {e}")
                    print(traceback.format_exc())
                    # Continue with next row
                    continue
                
                self.signals.update_progress.emit(current_row, total_rows)
            
            # Final save
            if processed_count > 0:
                try:
                    # Create DataFrame from all collected rows
                    self.output_df = pd.DataFrame(all_rows, columns=columns)
                    self.save_results()
                    print(f"Completed processing {processed_count} rows")
                except Exception as e:
                    print(f"Error in final save: {e}")
                    # Emergency CSV save
                    try:
                        csv_path = os.path.splitext(self.output_path)[0] + ".csv"
                        self.output_df.to_csv(csv_path, index=False)
                        print(f"Emergency save to CSV: {csv_path}")
                    except Exception as csv_error:
                        print(f"Emergency CSV save failed: {csv_error}")
            
            if self.running:
                self.signals.finished.emit()
        except Exception as e:
            error_message = str(e)
            print(f"Error in process_file: {error_message}")
            print(traceback.format_exc())
            self.signals.error.emit(error_message)
    
    def save_results(self):
        """Save the processed data to an Excel file"""
        if self.output_df is None or self.output_path is None:
            print("Cannot save: output_df or output_path is None")
            return
            
        try:
            # Create the directory if it doesn't exist
            os.makedirs(os.path.dirname(self.output_path), exist_ok=True)
            print(f"Saving to: {self.output_path}")
            
            # Check if the output path has the correct extension
            output_path = self.output_path
            if output_path.lower().endswith('.csv'):
                # Change to .xlsx for Excel export
                output_path = os.path.splitext(output_path)[0] + '.xlsx'
                print(f"Changed output path to Excel format: {output_path}")
            
            if self.output_df.empty:
                print("Warning: Output DataFrame is empty")
                
            # Ensure required columns exist in the correct order
            required_cols = ["Mfr Model", "Title", "Description", "Price"]
            for col in required_cols:
                if col not in self.output_df.columns:
                    self.output_df[col] = ""
            
            # Get all other columns
            all_columns = list(self.output_df.columns)
            other_cols = [col for col in all_columns if col not in required_cols]
            
            # Create a new DataFrame with the columns in the correct order
            ordered_cols = required_cols + other_cols
            ordered_df = pd.DataFrame(columns=ordered_cols)
            
            # Copy data from original DataFrame
            for col in ordered_cols:
                if col in self.output_df.columns:
                    ordered_df[col] = self.output_df[col]
            
            # Save to Excel file - without specifying engine
            print(f"Saving DataFrame with {len(ordered_df)} rows to {output_path}")
            ordered_df.to_excel(output_path, index=False)
            
            # Apply Excel formatting
            try:
                workbook = openpyxl.load_workbook(output_path)
                worksheet = workbook.active
                
                # Set row heights
                for row in worksheet.iter_rows():
                    worksheet.row_dimensions[row[0].row].height = 15
                
                # Set text wrapping for description
                for row in worksheet.iter_rows():
                    for cell in row:
                        col_name = worksheet.cell(row=1, column=cell.column).value
                        if col_name == "Description":
                            cell.alignment = Alignment(wrap_text=True)
                
                workbook.save(output_path)
                workbook.close()
                print(f"Excel formatting applied successfully to {output_path}")
            except Exception as e:
                print(f"Excel formatting error (non-critical): {e}")
                print(traceback.format_exc())
                
            print(f"File saved successfully: {output_path}")
            
            # If original path was CSV, save a CSV version too
            if self.output_path.lower().endswith('.csv'):
                try:
                    print(f"Also saving CSV version to: {self.output_path}")
                    ordered_df.to_csv(self.output_path, index=False)
                    print(f"CSV file saved: {self.output_path}")
                except Exception as csv_error:
                    print(f"Error saving CSV version: {csv_error}")
                    
        except Exception as e:
            print(f"Error saving results: {e}")
            print(traceback.format_exc())
            
            # Try saving as CSV as a fallback
            try:
                csv_path = os.path.splitext(self.output_path)[0] + ".csv"
                print(f"Attempting to save as CSV: {csv_path}")
                self.output_df.to_csv(csv_path, index=False)
                print(f"Saved as CSV: {csv_path}")
            except Exception as csv_error:
                print(f"Failed to save as CSV: {csv_error}")
                print(traceback.format_exc())


class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        try:
            self.authenticate_google_drive()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to authenticate with Google Drive: {str(e)}")
            raise
        self.setup_ui()
        self.add_row()
        self.processing_queue = []
        self.current_processing_index = -1
        self.error_count = 0
    
    def authenticate_google_drive(self):
        try:
            creds_path = os.path.expanduser("~/GoogleDriveMount/Web/zapier-454818-4e4abf368f57.json")
            scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
            creds = ServiceAccountCredentials.from_json_keyfile_name(creds_path, scope)
            self.gc = gspread.authorize(creds)
        except Exception as e:
            raise AppError(f"Google Drive authentication failed: {str(e)}")
    
    def get_drive_web_files(self):
        try:
            web_folder = os.path.expanduser("~/GoogleDriveMount/Web/")
            local_files = []
            if os.path.exists(web_folder):
                print(f"Looking for files in: {web_folder}")
                for filename in os.listdir(web_folder):
                    full_path = os.path.join(web_folder, filename)
                    if os.path.isfile(full_path):
                        if filename.endswith(('.csv', '.xlsx', '.xls')) and not filename.startswith('final_'):
                            local_files.append(filename)
                            print(f"Found file: {filename}")
            print(f"Found {len(local_files)} files in local Web folder")
            return sorted(local_files)
        except Exception as e:
            print(f"Error listing files: {e}")
            print(traceback.format_exc())
            return []
    
    def get_selected_files(self):
        selected_files = []
        try:
            for i in range(self.scroll_layout.count()):
                item = self.scroll_layout.itemAt(i)
                if not item:
                    continue
                row = item.widget()
                if row and isinstance(row, SheetRow) and hasattr(row, 'selected_file') and row.selected_file:
                    selected_files.append(row.selected_file)
        except Exception as e:
            print(f"Error getting selected files: {e}")
            print(traceback.format_exc())
        return selected_files
    
    def setup_ui(self):
        self.setWindowTitle("MK Processor 3.0.6")
        self.setGeometry(100, 100, 800, 600)
        self.setStyleSheet("""
            QWidget {
                background-color: #f0f0f0;
                font-family: Arial;
            }
            QLabel {
                color: #333333;
            }
            QLabel#headerLabel {
                color: #222222;
                font-size: 18px;
                font-weight: bold;
            }
            QLabel#statusLabel {
                color: #333333;
                font-size: 13px;
            }
            QPushButton {
                border-radius: 2px;
                padding: 5px 10px;
                font-weight: bold;
                min-height: 30px;
            }
            QPushButton#actionButton {
                background-color: #4285f4;
                color: white;
                border: none;
            }
            QPushButton#actionButton:hover {
                background-color: #3367d6;
            }
            QPushButton#actionButton:disabled {
                background-color: #a5c2f5;
            }
            QPushButton#secondaryButton {
                background-color: #f5f5f5;
                color: #333333;
                border: 1px solid #cccccc;
            }
            QPushButton#secondaryButton:hover {
                background-color: #e5e5e5;
            }
            QPushButton#dangerButton {
                background-color: #f5f5f5;
                color: #d32f2f;
                border: 1px solid #ffcdd2;
            }
            QPushButton#dangerButton:hover {
                background-color: #ffebee;
            }
            QScrollArea {
                border: none;
                background-color: transparent;
            }
            QScrollBar:vertical {
                border: none;
                background: #f5f5f5;
                width: 10px;
                border-radius: 5px;
            }
            QScrollBar::handle:vertical {
                background: #cccccc;
                min-height: 30px;
                border-radius: 5px;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                border: none;
                background: none;
            }
        """)
        
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(15)
        
        header = QLabel("MK Processor", self)
        header.setObjectName("headerLabel")
        header.setAlignment(Qt.AlignCenter)
        
        self.status_label = QLabel("Ready", self)
        self.status_label.setObjectName("statusLabel")
        self.status_label.setAlignment(Qt.AlignCenter)
        
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)
        
        self.start_all_btn = QPushButton("Start All", self)
        self.start_all_btn.setObjectName("actionButton")
        self.start_all_btn.clicked.connect(self.start_all)
        
        self.stop_all_btn = QPushButton("Stop All", self)
        self.stop_all_btn.setObjectName("secondaryButton")
        self.stop_all_btn.clicked.connect(self.stop_all)
        self.stop_all_btn.setEnabled(False)
        
        self.add_row_btn = QPushButton("Add File", self)
        self.add_row_btn.setObjectName("secondaryButton")
        self.add_row_btn.clicked.connect(self.add_row)
        
        self.clear_btn = QPushButton("Clear All", self)
        self.clear_btn.setObjectName("dangerButton")
        self.clear_btn.clicked.connect(self.clear_all)
        
        button_layout.addWidget(self.start_all_btn)
        button_layout.addWidget(self.stop_all_btn)
        button_layout.addWidget(self.add_row_btn)
        button_layout.addWidget(self.clear_btn)
        
        self.scroll_area = QScrollArea(self)
        self.scroll_area.setWidgetResizable(True)
        
        self.scroll_content = QWidget()
        self.scroll_layout = QVBoxLayout(self.scroll_content)
        self.scroll_layout.setAlignment(Qt.AlignTop)
        self.scroll_layout.setSpacing(10)
        self.scroll_layout.setContentsMargins(5, 5, 5, 5)
        
        self.scroll_area.setWidget(self.scroll_content)
        
        main_layout.addWidget(header)
        main_layout.addWidget(self.status_label)
        main_layout.addLayout(button_layout)
        main_layout.addWidget(self.scroll_area, 1)
    
    def add_row(self):
        try:
            row = SheetRow(len(self.scroll_layout), self)
            QApplication.processEvents()
            self.scroll_layout.addWidget(row)
            QApplication.processEvents()
            QTimer.singleShot(100, lambda: self.scroll_area.verticalScrollBar().setValue(
                self.scroll_area.verticalScrollBar().maximum()))
            QTimer.singleShot(500, self.refresh_all_rows)
        except Exception as e:
            print(f"Error adding row: {e}")
            print(traceback.format_exc())
            QMessageBox.warning(self, "Error", f"Error adding new row: {str(e)}")
    
    def refresh_all_rows(self):
        try:
            for i in range(self.scroll_layout.count()):
                item = self.scroll_layout.itemAt(i)
                if item:
                    row = item.widget()
                    if row and not row.running:
                        row.load_files()
        except Exception as e:
            print(f"Error refreshing rows: {e}")
    
    def clear_all(self):
        running_found = False
        for i in range(self.scroll_layout.count()):
            item = self.scroll_layout.itemAt(i)
            if item:
                row = item.widget()
                if row and row.running:
                    running_found = True
                    break
        if running_found:
            reply = QMessageBox.question(
                self, "Confirm",
                "Processing is currently running. Stop and clear all?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if reply != QMessageBox.Yes:
                return
        self.stop_all()
        while self.scroll_layout.count():
            item = self.scroll_layout.takeAt(0)
            if item and item.widget():
                item.widget().deleteLater()
        self.update_status("Ready")
        self.add_row()
    
    def start_all(self):
        valid_rows = []
        for i in range(self.scroll_layout.count()):
            item = self.scroll_layout.itemAt(i)
            if not item:
                continue
            row = item.widget()
            if row and row.file_dropdown.currentText() and row.prefix_input.text().strip():
                if hasattr(row, 'reset_state'):
                    row.reset_state()
                if hasattr(row, 'lock_controls'):
                    row.lock_controls(True)
                valid_rows.append(row)
        if not valid_rows:
            QMessageBox.warning(self, "Error", "Please add at least one file with a prefix")
            return
        self.start_all_btn.setEnabled(False)
        self.stop_all_btn.setEnabled(True)
        self.update_status("Starting sequential processing...")
        self.processing_queue = valid_rows
        self.current_processing_index = -1
        self.error_count = 0  # Reset error count
        QTimer.singleShot(100, self.process_next_row)
    
    def process_next_row(self):
        if not self.stop_all_btn.isEnabled():
            return
        self.current_processing_index += 1
        if self.current_processing_index >= len(self.processing_queue):
            if self.error_count > 0:
                self.update_status(f"All processing completed with {self.error_count} errors")
            else:
                self.update_status("All processing completed successfully")
            self.start_all_btn.setEnabled(True)
            self.stop_all_btn.setEnabled(False)
            return
        row = self.processing_queue[self.current_processing_index]
        file_info = row.get_selected_file()
        if file_info:
            self.update_status(f"Starting file {self.current_processing_index + 1} of {len(self.processing_queue)}: {file_info['name']}")
        self._start_row(row)
    
    def _start_row(self, row):
        if not row.running and not row.completed:
            row.start_processing()
    
    def stop_all(self):
        self.processing_queue = []
        for i in range(self.scroll_layout.count()):
            item = self.scroll_layout.itemAt(i)
            if not item:
                continue
            row = item.widget()
            if row:
                if row.running:
                    row.stop_processing()
                row.status_label.setText("Stopped")
                if hasattr(row, 'lock_controls'):
                    row.lock_controls(False)
        self.start_all_btn.setEnabled(True)
        self.stop_all_btn.setEnabled(False)
        self.update_status("Stopped")
    
    def update_status(self, message):
        self.status_label.setText(message)

def main():
    app = QApplication(sys.argv)
    try:
        print("Starting MK Processor v3.0.6")
        window = MainWindow()
        window.show()
        sys.exit(app.exec_())
    except Exception as e:
        print(f"Application error: {e}")
        print(traceback.format_exc())
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Critical)
        msg.setWindowTitle("Application Error")
        msg.setText(f"The application failed to start: {str(e)}")
        msg.setDetailedText(traceback.format_exc())
        msg.exec_()
        sys.exit(1)

if __name__ == "__main__":
    main()
